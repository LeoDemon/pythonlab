#  -*- coding: gbk -*-
# file_name: excel_test.py
# learning openpyxl for excel

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, StyleProxy, colors
from openpyxl import Workbook
from openpyxl import chart

def dealExcel():
    wb = load_workbook('../file/new_working_schedule.xlsx')
    for irow, sheetName in enumerate(wb.get_sheet_names()):
        print('The %d sheet name: [%s]' % (irow, sheetName))

    print('\nGet a existing sheet...')
    ft1 = Font(color=colors.RED, bold=True, size=12)
    al1 = Alignment(vertical='center', horizontal='center')
    bd1 = Border(left=Side(border_style='thin', color='00000000'), right=Side(border_style='thin', color='00000000'),
                 top=Side(border_style='thin', color='00000000'), bottom=Side(border_style='thin', color='00000000'))

    ws1 = wb['test_plan']
    all_b = 'B'+str(ws1.max_row+1)
    max_b = 'B'+str(ws1.max_row)
    sum_b = '=SUM(B2:%s)'%max_b
    ws1[all_b] = sum_b
    print('max_row==',ws1.max_row,',max_column==',ws1.max_column)
    for icn in range(ws1.max_column):
        new_col = ord('A') + icn
        print('%s-%d-%s'%(chr(new_col)+'1',len(ws1[chr(new_col)+'1'].value),ws1[chr(new_col)+'1'].value))
        ws1.column_dimensions[chr(new_col)].width = len(ws1[chr(new_col)+'1'].value) + 10

    for ir, row in enumerate(ws1):
        ws1.row_dimensions[ir+1].height = 36    #row height
        if 0 == ir:
            for cell in row:
                cell.alignment = al1
                cell.font = ft1
                cell.border = bd1
        else:
            for cell in row:
                cell.alignment = al1
                cell.border = bd1

    print('\nCreate a new sheet...')
    ws2 = wb.create_sheet('test_schedule', 3)
    ws2.sheet_properties.tabColor = '107200'
    for irow, sheet in enumerate(wb):
        print('The %d sheet name: [%s]' % (irow, sheet.title))

    print('\nEvaluate for new sheet...')
    ws2.freeze_panes = 'A2'    #freeze the 1st row
    for ir in range(1,6):
        if 1 == ir:
            for ic in range(1,6):
                ws2.cell(row=ir, column=ic, value='id%d'%ic)
                ws2.cell(row=ir, column=ic).alignment = al1
                ws2.cell(row=ir, column=ic).border = bd1
                ws2.cell(row=ir, column=ic).font = ft1
        else:
            for ic in range(1,6):
                ws2.cell(row=ir, column=ic, value=ir+ic)
                ws2.cell(row=ir, column=ic).alignment = al1
                ws2.cell(row=ir, column=ic).border = bd1

    #drawing chart
    refObj = chart.Reference(ws2, min_col=1, min_row=1, max_col=5, max_row=5)
    seriesObj = chart.Series(refObj, title='First-Series')
    chartObj = chart.BarChart()
    chartObj.title = 'First-Chart'
    chartObj.append(seriesObj)
    ws2.add_chart(chartObj, 'F10')

    print('\nRemove the sheet[dev-schedule]')
    wb.remove_sheet(wb.get_sheet_by_name('dev-schedule'))

    print('\nSave file...')
    wb.save('../file/working_schedule.xlsx')

if "__main__" == __name__:
    dealExcel()
else:
    print("being imported by another module...")

